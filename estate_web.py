import re
import openpyxl
import requests
from bs4 import BeautifulSoup


url = "https://www.home.ge/saxlebi-agarakebi/iyideba-saxlebi-agarakebi/iyideba-sakhli-bakuriani-297726.html"

# html convert to text to be readable
html_text = requests.get(url).text

# parse the code and we got the raw html
soup = BeautifulSoup(html_text, 'html.parser')

# try to find all img tag
images = soup.find_all('img')

# make a dictionary and value is list, because append be usable

link_list = {'img':[]}

# I will loop all img tag
for image in images:
    # find src attr in img tag
    link = image.get('src')
    # from inspect i Know the last keywords of this photo, I will see and use endswith to be valid
    if link.endswith("large.jpg"):
        #restul will add in link_list
        link_list['img'].append(link)

# this is for title to see h1
h1 = soup.find_all('h1')

title = []
for char in h1:
    title.append(char.text)

price = []
price_tag = soup.find_all('div', id='df_field_price')

for tag in price_tag:
    price.append(tag.text.lstrip())



# search div class value

sqm_tag = soup.find_all('div', class_= "value")
# add empty lsit for appending
sqm_list = []

#loop the list
for sqm in sqm_tag:
    # while this is written by human when the info comes from web i try to delete all right and left space
    sqm_info = sqm.text.lstrip().rstrip()
    #while i have many div with the same class name, i know how many words is written in sqm value, i split and got 2 length of list
    sqm_split = sqm_info.split(" ")
    # make a validation and reduce my search area
    if len(sqm_split) == 2:
        #i know from the web that house sqm is 113, so the one of the list must be contain 113
        if "113" in sqm_split:
            # then i convert the list in str to append and save in the list
            sqm_join = "".join(sqm_split)
            #i add in sqm_list 
            sqm_list.append(sqm_join)
            
            

#Create a new Excel workbook
workbook = openpyxl.Workbook()
worksheet = workbook.active

#Write the title to cell A1
worksheet.cell(row=1, column=1, value="Image Links")
worksheet.cell(row=1, column =2, value = "Title")
worksheet.cell(row = 1, column = 3, value = "Sqm")
worksheet.cell(row = 1, column = 4, value = "Price")

#Write the links to the worksheet

for i, link in enumerate(link_list['img']):
    worksheet.cell(row = i+2, column = 1, value = link)
    print(link)

for i, t in enumerate(title):
    worksheet.cell(row = i+2, column = 2, value = t)

for i, p in enumerate(price):
    worksheet.cell(row = i+2, column = 4, value = p )

for i, sqm in enumerate(sqm_list):
    worksheet.cell(row = i+2, column = 3, value = sqm )

#save the workbook

workbook.save('links.xlsx')



